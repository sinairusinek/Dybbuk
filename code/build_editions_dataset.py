"""Build the integrated editions dataset.

Inputs:
  - data/editions.csv                        (our 16 editions, hand-maintained)
  - edition metadata/DybbukCatalogue May2024.xlsx  (DiJeSt-style catalogue)
  - edition metadata/PerformanceEvents_Report_*.xlsx  (DB performance events, latest by mtime)

Outputs:
  - data/editions.json       nested per-edition records (for XML header injection)
  - data/editions_flat.csv   flat one-row-per-edition slice (for visualization)

Join cascade:
  edition.transkribus_doc_id
    → TranskribusPlayStatus.TranskribusDocId    → Play key (canonical play name)
    → Lateiner Plays.English Name / Yiddish Name → Expression ID (DiJeSt expression)
    → fan-out:
        Lateiner hafakot.expression / Play KEY  → productions
        ProfessionalRoles- Lateiner.Play         → roles
        songs_Lateiner and Hurwitz.Play Key      → songs
        Score-print-editions.work id             → print-edition catalogue rows
        documentsitems.work id                   → document items
"""
from __future__ import annotations

import csv
import datetime as _dt
import json
import re
from pathlib import Path

import openpyxl


def _json_default(o):
    if isinstance(o, (_dt.datetime, _dt.date)):
        return o.isoformat()
    raise TypeError(f"not serializable: {type(o).__name__}")


ROOT = Path(__file__).resolve().parent.parent
EDITIONS_CSV = ROOT / "data" / "editions.csv"
XLSX = ROOT / "edition metadata" / "DybbukCatalogue May2024.xlsx"
PERF_DIR = ROOT / "edition metadata"
OUT_JSON = ROOT / "data" / "editions.json"
OUT_CSV = ROOT / "data" / "editions_flat.csv"
OUT_MD = ROOT / "data" / "editions.md"


def _latest_perf_events_file() -> Path | None:
    candidates = sorted(PERF_DIR.glob("PerformanceEvents_Report_*.xlsx"))
    return candidates[-1] if candidates else None


def _year_from_date(s) -> int | None:
    if not s:
        return None
    if isinstance(s, (_dt.datetime, _dt.date)):
        return s.year
    m = re.search(r"\b(18|19|20)\d{2}\b", str(s))
    return int(m.group()) if m else None


def _norm_venue(v) -> str:
    return re.sub(r"\s+", " ", (v or "").strip().lower())


def build_performance_events(perf_path: Path, productions_by_eid_key: dict) -> dict:
    """Read the DB PerformanceEvents report and merge each row with the
    matching hafakot production (by year/type) to add `venue_alt` etc.

    Returns dict keyed by (lowercased roman title) -> list[event dict].
    """
    wb = openpyxl.load_workbook(perf_path, data_only=True)
    ws = wb["PerformanceEvents"]
    H = [c.value for c in ws[1]]
    out: dict[str, list] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        rec = {h: v for h, v in zip(H, r) if h}
        if not any(rec.values()):
            continue
        rom = (rec.get("Roman Title") or "").strip()
        yid = (rec.get("Expression") or "").strip()
        if not rom and not yid:
            continue
        event = {
            "id": rec.get("Id"),
            "event_type": rec.get("Event Type"),
            "date": rec.get("Date").isoformat() if isinstance(rec.get("Date"), (_dt.datetime, _dt.date)) else rec.get("Date"),
            "year": _year_from_date(rec.get("Date")),
            "venue": (rec.get("Held at") or "") or None,
            "actor_character": (rec.get("Actor (Character)") or "") or None,
            "person_role": (rec.get("Person (Professional Role)") or "") or None,
            "roman_title": rom or None,
            "yiddish_title": yid or None,
        }
        key = rom.lower() if rom else None
        if key:
            out.setdefault(key, []).append(event)
    return out


def _match_events_to_productions(events: list[dict], productions: list[dict]) -> list[dict]:
    """Augment each event with venue_alt/source from a matching hafakot row.

    Match heuristic: same (year, event_type). For unique single-event plays,
    fall back to position match.
    """
    if not events:
        return events
    used = set()
    for ev in events:
        ev_year = ev.get("year")
        ev_type = (ev.get("event_type") or "").lower()
        match = None
        for i, p in enumerate(productions):
            if i in used:
                continue
            p_year = p.get("Year")
            p_year_int = int(p_year) if isinstance(p_year, (int, float)) else None
            p_type = (p.get("Type") or "").lower()
            if p_year_int == ev_year and (not ev_type or not p_type or ev_type == p_type):
                match = (i, p)
                break
        # Fallback: single-event + single-production play
        if match is None and len(events) == 1 and len(productions) == 1:
            match = (0, productions[0])
        if match is not None:
            i, p = match
            used.add(i)
            haf_venue = p.get("Theatre") or p.get("PremierePlace")
            if haf_venue and _norm_venue(haf_venue) != _norm_venue(ev.get("venue")):
                ev["venue_alt"] = haf_venue
            ev["source_catalogue"] = p.get("source") or p.get("dating Source")
            if p.get("Notes"):
                ev["notes"] = p["Notes"]
    return events


def to_int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def load_sheet(wb, name):
    """Return (headers, list[dict])."""
    ws = wb[name]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append({h: v for h, v in zip(headers, r) if h is not None})
    return headers, rows


def _write_markdown(enriched: list[dict], out_path: Path) -> None:
    """Write a human-readable snapshot of the corpus to data/editions.md."""
    today = _dt.date.today().isoformat()
    lines: list[str] = []
    lines.append("# Editions corpus — live snapshot")
    lines.append("")
    lines.append(f"*Auto-generated by `code/build_editions_dataset.py` on {today}. "
                 "Do not edit by hand — edit `data/editions.csv` and re-run.*")
    lines.append("")
    lines.append(f"**{len(enriched)} editions** "
                 f"({sum(1 for e in enriched if e.get('expression_id'))} with expression IDs)")
    lines.append("")

    # 1. Overview table — includes publication info
    lines.append("## Overview")
    lines.append("")
    lines.append("| Title | Author | Year | Pub. place | Publisher | Pub. place (of publisher) "
                 "| Printer | Printer place | docId | eid | HTR | Folder |")
    lines.append("|---|---|---|---|---|---|---|---|---|---|---|---|")
    for e in enriched:
        yp = e.get("year_printed") or e.get("year_written") or "—"
        eid = e.get("expression_id") or "—"
        ready = e.get("transkribus_ready") or "—"
        lines.append(
            f"| {e.get('title') or ''} "
            f"| {e.get('author') or ''} "
            f"| {yp} "
            f"| {e.get('publication_place') or ''} "
            f"| {e.get('publisher') or ''} "
            f"| {e.get('publisher_place') or ''} "
            f"| {e.get('printer') or ''} "
            f"| {e.get('printer_place') or ''} "
            f"| {e.get('transkribus_doc_id') or ''} "
            f"| {eid} "
            f"| {ready} "
            f"| `{e.get('folder') or ''}` |"
        )
    lines.append("")

    # 3. Vocalization conventions (RA-filled)
    lines.append("## Vocalization conventions (RA)")
    lines.append("")
    lines.append("| Title | rafe | voc. position | speakers/stage vocalized | notes |")
    lines.append("|---|---|---|---|---|")
    for e in enriched:
        if not any(e.get(k) for k in
                   ("rafe", "vocalization_position", "speakers_stage_vocalized")):
            continue
        lines.append(
            f"| {e.get('title')} "
            f"| {e.get('rafe') or ''} "
            f"| {e.get('vocalization_position') or ''} "
            f"| {e.get('speakers_stage_vocalized') or ''} "
            f"| {e.get('notes') or ''} |"
        )
    lines.append("")

    # 4. Catalogue enrichment counts
    lines.append("## Catalogue enrichment (per edition)")
    lines.append("")
    lines.append("| Title | eid | catalogue title | tags | productions | events (P/S) | songs | roles |")
    lines.append("|---|---|---|---|---|---|---|---|")
    for e in enriched:
        exp = e.get("expression") or {}
        evs = e.get("performance_events") or []
        prem = sum(1 for x in evs if (x.get("event_type") or "").lower() == "premiere")
        show = sum(1 for x in evs if (x.get("event_type") or "").lower() == "show")
        lines.append(
            f"| {e.get('title')} "
            f"| {e.get('expression_id') or '—'} "
            f"| {exp.get('english_name') or ''} "
            f"| {exp.get('tags') or ''} "
            f"| {len(e.get('productions') or [])} "
            f"| {len(evs)} ({prem}/{show}) "
            f"| {len(e.get('songs') or [])} "
            f"| {len(e.get('roles') or [])} |"
        )
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    print("Loading editions.csv …")
    with EDITIONS_CSV.open(encoding="utf-8") as f:
        editions = list(csv.DictReader(f))
    print(f"  {len(editions)} editions")

    print("Loading xlsx …")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    _, tps = load_sheet(wb, "TranskribusPlayStatus")
    _, lateiner = load_sheet(wb, "Lateiner Plays")
    _, hafakot = load_sheet(wb, "Lateiner hafakot")
    _, roles = load_sheet(wb, "ProfessionalRoles- Lateiner")
    _, songs = load_sheet(wb, "songs_Lateiner and Hurwitz")
    _, scores = load_sheet(wb, "Score-print-editions")
    _, docitems = load_sheet(wb, "documentsitems")
    print(f"  TranskribusPlayStatus={len(tps)} LateinerPlays={len(lateiner)} "
          f"hafakot={len(hafakot)} roles={len(roles)} songs={len(songs)} "
          f"scores={len(scores)} docitems={len(docitems)}")

    # docId → tps row
    tps_by_doc = {to_int(r.get("TranskribusDocId")): r
                  for r in tps if to_int(r.get("TranskribusDocId"))}

    # name → expression_id (try both English and Yiddish names)
    name_to_eid = {}
    eid_to_lateiner = {}
    for r in lateiner:
        eid = to_int(r.get("Expression ID"))
        if eid is None:
            continue
        eid_to_lateiner[eid] = r
        for key in ("English Name", "Yiddish Name"):
            v = (r.get(key) or "").strip()
            if v:
                name_to_eid[v.lower()] = eid

    # Manual overrides for Transkribus 'Play' values that don't string-match exactly
    play_to_eid_overrides = {
        "yidele, oder der emes un der sheker": 3927,  # → "Yidele oder der emes un sheyker"
    }
    # Direct docId → expression_id overrides (for editions absent from TranskribusPlayStatus)
    doc_to_eid_overrides = {
        828424: 3867,  # Das Yudishe Kind → "Dos yidishe kind"
        820937: 3884,  # Isha Raa → "Ishe roeh" / אישה רע
    }

    # Fan-out indexes keyed by expression_id OR play-name string
    def index_by(rows, *keys):
        idx_eid: dict[int, list] = {}
        idx_name: dict[str, list] = {}
        for r in rows:
            eid = None
            for k in keys:
                v = r.get(k)
                ev = to_int(v)
                if ev is not None:
                    eid = ev
                    break
            if eid is not None:
                idx_eid.setdefault(eid, []).append(r)
            for k in keys:
                v = r.get(k)
                if isinstance(v, str) and v.strip():
                    idx_name.setdefault(v.strip().lower(), []).append(r)
        return idx_eid, idx_name

    perf_path = _latest_perf_events_file()
    events_by_roman: dict[str, list] = {}
    if perf_path:
        events_by_roman = build_performance_events(perf_path, {})
        print(f"  PerformanceEvents: {sum(len(v) for v in events_by_roman.values())} rows "
              f"across {len(events_by_roman)} titles ({perf_path.name})")
    else:
        print("  PerformanceEvents: no report file found")

    hafakot_eid, hafakot_name = index_by(hafakot, "expression", "Play KEY")
    roles_eid, roles_name = index_by(roles, "PersonKey", "Play")
    songs_eid, songs_name = index_by(songs, "Play Key")
    scores_eid, _ = index_by(scores, "work id")
    docitems_eid, _ = index_by(docitems, "work id")

    enriched: list[dict] = []
    unmatched_no_eid: list[str] = []

    for ed in editions:
        doc_id = to_int(ed.get("transkribus_doc_id"))
        rec: dict = {
            # our authored columns
            **ed,
            "transkribus_doc_id": doc_id,
            "transkribus_collection_id": to_int(ed.get("transkribus_collection_id")),
        }
        rec["year_written"] = to_int(ed.get("year_written")) or ed.get("year_written") or None
        rec["year_printed"] = to_int(ed.get("year_printed")) or ed.get("year_printed") or None

        # tps row
        tps_row = tps_by_doc.get(doc_id) if doc_id else None
        play_key = (tps_row.get("Play") if tps_row else None) or ""
        rec["catalogue_play_key"] = play_key or None
        rec["catalogue_yiddish_name"] = tps_row.get("name") if tps_row else None
        rec["transkribus_status"] = {
            "layout": tps_row.get("layout") if tps_row else None,
            "script": tps_row.get("script") if tps_row else None,
            "all_corrected": tps_row.get("ALLcorrected?") if tps_row else None,
            "all_proofed": tps_row.get("ALLproofed?") if tps_row else None,
            "ready_for_encoding": tps_row.get("ready for encoding") if tps_row else None,
        } if tps_row else None

        # expression_id resolution
        eid = doc_to_eid_overrides.get(doc_id)
        if eid is None and play_key:
            k = play_key.strip().lower()
            eid = play_to_eid_overrides.get(k) or name_to_eid.get(k)
        if eid is None:
            # last-chance fallback by our own English title
            eid = name_to_eid.get((ed.get("title") or "").strip().lower())
        rec["expression_id"] = eid

        if eid:
            lp = eid_to_lateiner.get(eid, {})
            rec["expression"] = {
                "expression_id": eid,
                "english_name": lp.get("English Name"),
                "yiddish_name": lp.get("Yiddish Name"),
                "tags": lp.get("TAGS"),
                "genre": lp.get("Genre"),
                "certainty": lp.get("certainty"),
                "comments": lp.get("comments"),
                "author_id": to_int(lp.get("author")),
                "expression_note": lp.get("expression"),
                "do_we_have_a_copy": lp.get("do we have a copy"),
            }
            rec["productions"] = [
                {k: v for k, v in r.items() if v is not None}
                for r in (hafakot_eid.get(eid, [])
                          + hafakot_name.get(play_key.strip().lower(), []))
            ]
            rec["roles"] = [
                {k: v for k, v in r.items() if v is not None}
                for r in roles_name.get(play_key.strip().lower(), [])
            ]
            rec["songs"] = [
                {k: v for k, v in r.items() if v is not None}
                for r in songs_name.get(play_key.strip().lower(), [])
            ]
            rec["print_edition_catalogue"] = [
                {k: v for k, v in r.items() if v is not None}
                for r in scores_eid.get(eid, [])
            ]
            rec["document_items"] = [
                {k: v for k, v in r.items() if v is not None}
                for r in docitems_eid.get(eid, [])
            ]

            # Performance events — DB report is canonical; merge venue_alt from hafakot
            en = (lp.get("English Name") or "").strip().lower()
            stem = en.split(",")[0].split(" oder ")[0].strip() if en else ""
            evs = (events_by_roman.get(en)
                   or events_by_roman.get(stem)
                   or [])
            rec["performance_events"] = _match_events_to_productions(
                [dict(e) for e in evs], rec["productions"]
            )
        else:
            unmatched_no_eid.append(f"{ed.get('title')} (doc {doc_id})")
            rec["expression"] = None
            rec["productions"] = []
            rec["roles"] = []
            rec["songs"] = []
            rec["print_edition_catalogue"] = []
            rec["document_items"] = []
            rec["performance_events"] = []

        enriched.append(rec)

    # Write JSON
    OUT_JSON.write_text(
        json.dumps({"editions": enriched}, ensure_ascii=False, indent=2,
                   default=_json_default),
        encoding="utf-8",
    )
    print(f"Wrote {OUT_JSON}")

    # Write flat CSV
    flat_cols = [
        "title", "author", "year_written", "year_printed",
        "publication_place", "publisher", "publisher_place",
        "printer", "printer_place",
        "transkribus_doc_id", "transkribus_collection_id", "transkribus_url",
        "transkribus_ready", "folder",
        "catalogue_play_key", "catalogue_yiddish_name",
        "expression_id",
        "expression_english_name", "expression_yiddish_name",
        "tags", "genre", "certainty",
        "n_productions", "n_performance_events", "n_premieres", "n_shows",
        "n_roles", "n_songs",
        "n_print_edition_catalogue", "n_document_items",
        "rafe", "vocalization_position", "speakers_stage_vocalized", "notes",
    ]
    with OUT_CSV.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=flat_cols)
        w.writeheader()
        for r in enriched:
            exp = r.get("expression") or {}
            w.writerow({
                "title": r.get("title"),
                "author": r.get("author"),
                "year_written": r.get("year_written"),
                "year_printed": r.get("year_printed"),
                "publication_place": r.get("publication_place"),
                "publisher": r.get("publisher"),
                "publisher_place": r.get("publisher_place"),
                "printer": r.get("printer"),
                "printer_place": r.get("printer_place"),
                "transkribus_doc_id": r.get("transkribus_doc_id"),
                "transkribus_collection_id": r.get("transkribus_collection_id"),
                "transkribus_url": r.get("transkribus_url"),
                "transkribus_ready": r.get("transkribus_ready"),
                "folder": r.get("folder"),
                "catalogue_play_key": r.get("catalogue_play_key"),
                "catalogue_yiddish_name": r.get("catalogue_yiddish_name"),
                "expression_id": r.get("expression_id"),
                "expression_english_name": exp.get("english_name"),
                "expression_yiddish_name": exp.get("yiddish_name"),
                "tags": exp.get("tags"),
                "genre": exp.get("genre"),
                "certainty": exp.get("certainty"),
                "n_productions": len(r.get("productions") or []),
                "n_performance_events": len(r.get("performance_events") or []),
                "n_premieres": sum(1 for e in (r.get("performance_events") or [])
                                   if (e.get("event_type") or "").lower() == "premiere"),
                "n_shows": sum(1 for e in (r.get("performance_events") or [])
                               if (e.get("event_type") or "").lower() == "show"),
                "n_roles": len(r.get("roles") or []),
                "n_songs": len(r.get("songs") or []),
                "n_print_edition_catalogue": len(r.get("print_edition_catalogue") or []),
                "n_document_items": len(r.get("document_items") or []),
                "rafe": r.get("rafe"),
                "vocalization_position": r.get("vocalization_position"),
                "speakers_stage_vocalized": r.get("speakers_stage_vocalized"),
                "notes": r.get("notes"),
            })
    print(f"Wrote {OUT_CSV}")

    # Write human-readable markdown snapshot
    _write_markdown(enriched, OUT_MD)
    print(f"Wrote {OUT_MD}")

    print(f"\n{len(enriched)} editions enriched")
    print(f"  expression_id resolved: {sum(1 for r in enriched if r['expression_id'])}")
    if unmatched_no_eid:
        print("  UNMATCHED (no expression_id):")
        for u in unmatched_no_eid:
            print(f"    - {u}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
