"""Export a speaker-who review workbook for the RA (Noa).

Scans every play's page_annotated/ for `speaker {...}` spans, and builds an
.xlsx (Google-Sheets-convertible) with:

  * "Missing who"  — one row per speaker span with no xmlid, with a per-play
                     role dropdown to pick the right who.
  * "All mappings" — one row per unique (play, printed label, current who)
                     combination with occurrence counts, so the whole corpus
                     can be reviewed without 12k rows. Dropdown = correction;
                     empty = current who confirmed.
  * "Roles"        — hidden sheet holding the per-play dropdown lists
                     (cast_dict roles + any xmlid actually used in the play).

Deep links point at https://app.transkribus.org/collection/COL/doc/DOC/detail/PAGE.
`line_id` columns are the stable keys for applying decisions back.

Run:  python3.11 code/annotation/export_speaker_who_review.py
Writes: data/review/speaker_who_review_<date>.xlsx (+ .tsv occurrence dump)
"""
from __future__ import annotations

import csv
import datetime as dt
import json
import re
import sys
import xml.etree.ElementTree as ET
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]  # YiDraCor/
DATA = ROOT / "data"
COL = 2372172
DEEPLINK = "https://app.transkribus.org/collection/{col}/doc/{doc}/detail/{page}"

SPEAKER_RE = re.compile(r"speaker\s*\{([^}]*)\}")
ATTR_RE = re.compile(r"(\w+)\s*:\s*([^;]*);")

SPECIAL_OPTIONS = [
    "?? unsure — see note",
    "XX not a speaker",
    "++ new role — see note",
]

# --- trivial-mapping detection (label obviously names its assigned role) ----
import unicodedata

FINALS = str.maketrans("ךםןףץ", "כמנפצ")
# crude Yiddish -> Latin consonant skeleton, for comparing labels to xmlids
YID2LAT = {
    "א": "", "ב": "b", "ג": "g", "ד": "d", "ה": "h", "ו": "", "ז": "z",
    "ח": "h", "ט": "t", "י": "", "כ": "k", "ל": "l", "מ": "m", "נ": "n",
    "ס": "s", "ע": "", "פ": "p", "צ": "ts", "ק": "k", "ר": "r", "ש": "s",
    "ת": "t",
}


def norm_yid(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if not (0x0591 <= ord(c) <= 0x05C7))
    s = re.sub(r"[^א-תA-Za-z0-9 ]", "", s).translate(FINALS)
    return re.sub(r"\s+", " ", s).strip()


def skeleton_yid(s: str) -> str:
    return "".join(YID2LAT.get(c, "") for c in norm_yid(s))


def skeleton_lat(s: str) -> str:
    s = re.sub(r"[^a-z]", " ", s.lower())
    s = re.sub(r"[aeiouwyj]", "", s)
    s = s.replace("sh", "s").replace("kh", "h").replace("c", "k")
    return s.replace(" ", "")


def edit1(a: str, b: str) -> bool:
    if abs(len(a) - len(b)) > 1:
        return False
    if a == b:
        return True
    for i in range(min(len(a), len(b))):
        if a[i] != b[i]:
            return a[i + 1:] == b[i + 1:] or a[i:] == b[i + 1:] or a[i + 1:] == b[i:]
    return True  # differ only by one trailing char


def is_trivial(label: str, xmlid: str, role: dict) -> bool:
    L = norm_yid(label)
    if not L:
        return False
    cands = set()
    for v in [role.get("form", ""), role.get("bare", "")] + list(role.get("prefix_variants", [])):
        v = norm_yid(v)
        if v:
            cands.add(v)
            cands.update(t for t in v.split() if len(t) > 2)
    for c in cands:
        if L == c or (len(L) > 2 and (c.startswith(L) or L.startswith(c))):
            return True
        if len(L) > 3 and edit1(L, c):
            return True
    sk = skeleton_yid(label)
    if sk:
        for part in [xmlid] + xmlid.split("_"):
            skx = skeleton_lat(part)
            if not skx:
                continue
            # short skeletons carry too little signal for fuzzy matching
            if sk == skx or (max(len(sk), len(skx)) >= 3 and edit1(sk, skx)):
                return True
    return False


def load_doc_ids() -> dict[str, int]:
    docs = {}
    with open(DATA / "editions.csv", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            folder = (row.get("folder") or "").strip()
            doc = (row.get("transkribus_doc_id") or "").strip()
            if folder and doc.isdigit():
                docs[folder] = int(doc)
    return docs


def parse_page(path: Path):
    """Yield (line_id, line_text, offset, length, xmlid_or_empty) per speaker span."""
    tree = ET.parse(path)
    for tl in tree.getroot().iter():
        if not tl.tag.endswith("}TextLine"):
            continue
        custom = tl.get("custom") or ""
        if "speaker" not in custom:
            continue
        text = ""
        for te in tl:
            if te.tag.endswith("}TextEquiv"):
                for u in te:
                    if u.tag.endswith("}Unicode") and u.text:
                        text = u.text
                break
        for m in SPEAKER_RE.finditer(custom):
            attrs = dict(ATTR_RE.findall(m.group(1)))
            try:
                off, ln = int(attrs.get("offset", 0)), int(attrs.get("length", 0))
            except ValueError:
                off, ln = 0, 0
            label = text[off : off + ln].strip()
            yield tl.get("id") or "", text, label, attrs.get("xmlid", "").strip()


def collect():
    docs = load_doc_ids()
    occurrences = []  # dicts
    roles_by_play = {}  # play -> {xmlid: form}
    for play_dir in sorted(DATA.iterdir()):
        pa = play_dir / "page_annotated"
        if not pa.is_dir():
            continue
        play = play_dir.name
        if play not in docs:
            print(f"WARN: no transkribus_doc_id for folder {play!r} — links will be blank", file=sys.stderr)
        cast = {}
        cd_path = play_dir / "cast_dict.json"
        if cd_path.exists():
            cd = json.loads(cd_path.read_text(encoding="utf-8"))
            for xid, info in (cd.get("roles") or {}).items():
                cast[xid] = info.get("form") or info.get("bare") or xid
        roles_by_play[play] = cast
        for f in sorted(pa.glob("*.xml")):
            page_nr = int(f.name.split("_")[0])
            for line_id, line_text, label, xmlid in parse_page(f):
                occurrences.append(
                    dict(
                        play=play,
                        doc=docs.get(play, ""),
                        page=page_nr,
                        file=f.name,
                        line_id=line_id,
                        line_text=line_text,
                        label=label,
                        xmlid=xmlid,
                    )
                )
                if xmlid and xmlid not in cast:
                    cast[xmlid] = xmlid  # used in play but absent from cast_dict
    return occurrences, roles_by_play


def role_option(xmlid: str, form: str) -> str:
    return f"{form} ({xmlid})" if form and form != xmlid else xmlid


def build_workbook(occurrences, roles_by_play, out_xlsx: Path):
    wb = Workbook()
    hdr_font = Font(bold=True)
    hdr_fill = PatternFill("solid", fgColor="DDDDDD")
    miss_fill = PatternFill("solid", fgColor="FFF2CC")

    # ---- hidden Roles sheet + named ranges -------------------------------
    roles_ws = wb.create_sheet("Roles")
    play_range = {}
    for ci, play in enumerate(sorted(roles_by_play), start=1):
        col = get_column_letter(ci)
        roles_ws.cell(row=1, column=ci, value=play)
        opts = [role_option(x, f) for x, f in sorted(roles_by_play[play].items())]
        opts += SPECIAL_OPTIONS
        for ri, opt in enumerate(opts, start=2):
            roles_ws.cell(row=ri, column=ci, value=opt)
        name = f"Roles_{ci}"
        dn = DefinedName(name, attr_text=f"Roles!${col}$2:${col}${1 + len(opts)}")
        if hasattr(wb.defined_names, "append"):  # openpyxl < 3.1
            wb.defined_names.append(dn)
        else:
            wb.defined_names.add(dn)
        play_range[play] = name
    roles_ws.sheet_state = "hidden"

    def add_dv(ws, col_letter, rows_by_play):
        """rows_by_play: play -> list of row numbers (contiguous runs collapsed)."""
        for play, rows in rows_by_play.items():
            dv = DataValidation(
                type="list", formula1=f"={play_range[play]}", allow_blank=True,
                showDropDown=False,
            )
            ws.add_data_validation(dv)
            rows = sorted(rows)
            start = prev = rows[0]
            for r in rows[1:] + [None]:
                if r != (prev or 0) + 1:
                    dv.add(f"{col_letter}{start}:{col_letter}{prev}")
                    start = r
                prev = r

    def write_header(ws, headers, widths):
        for ci, (h, w) in enumerate(zip(headers, widths), start=1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font, c.fill = hdr_font, hdr_fill
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"

    def link_formula(o):
        if not o["doc"]:
            return ""
        url = DEEPLINK.format(col=COL, doc=o["doc"], page=o["page"])
        return f'=HYPERLINK("{url}", "p.{o["page"]}")'

    # ---- Missing who ------------------------------------------------------
    ws = wb.active
    ws.title = "Missing who"
    write_header(
        ws,
        ["Play", "Page", "Link", "Speaker label", "Full line", "Choose who ▼", "Notes", "line_id", "file"],
        [28, 6, 8, 18, 55, 26, 30, 14, 22],
    )
    rows_by_play = defaultdict(list)
    r = 2
    for o in [o for o in occurrences if not o["xmlid"]]:
        ws.cell(row=r, column=1, value=o["play"])
        ws.cell(row=r, column=2, value=o["page"])
        ws.cell(row=r, column=3, value=link_formula(o))
        ws.cell(row=r, column=4, value=o["label"]).fill = miss_fill
        ws.cell(row=r, column=5, value=o["line_text"])
        ws.cell(row=r, column=8, value=o["line_id"])
        ws.cell(row=r, column=9, value=o["file"])
        rows_by_play[o["play"]].append(r)
        r += 1
    add_dv(ws, "F", rows_by_play)

    # ---- Mappings to check ------------------------------------------------
    agg = defaultdict(lambda: {"n": 0, "pages": set(), "first": None})
    for o in occurrences:
        if not o["xmlid"]:
            continue
        k = (o["play"], o["label"], o["xmlid"])
        a = agg[k]
        a["n"] += 1
        a["pages"].add(o["page"])
        if a["first"] is None:
            a["first"] = o
    cast_dicts = {}
    for play in roles_by_play:
        p = DATA / play / "cast_dict.json"
        cast_dicts[play] = (json.loads(p.read_text(encoding="utf-8")).get("roles") or {}) if p.exists() else {}
    check, trivial = [], []
    for (play, label, xmlid), a in agg.items():
        role = cast_dicts[play].get(xmlid, {})
        (trivial if is_trivial(label, xmlid, role) else check).append((play, label, xmlid, a))

    ws2 = wb.create_sheet("Mappings to check")
    write_header(
        ws2,
        ["Play", "Speaker label (printed)", "Current who", "Count", "Example link",
         "Correction ▼ (empty = OK)", "Notes", "doc", "page"],
        [28, 20, 24, 7, 12, 26, 30, 10, 7],
    )
    # base URL parked in J1; per-row link = HYPERLINK($J$1 & doc & "/detail/" & page)
    ws2["J1"] = f"https://app.transkribus.org/collection/{COL}/doc/"
    rows_by_play = defaultdict(list)
    r = 2
    for play, label, xmlid, a in sorted(check, key=lambda t: (t[0], -t[3]["n"], t[1])):
        form = roles_by_play[play].get(xmlid, xmlid)
        ws2.cell(row=r, column=1, value=play)
        ws2.cell(row=r, column=2, value=label)
        ws2.cell(row=r, column=3, value=role_option(xmlid, form))
        ws2.cell(row=r, column=4, value=a["n"])
        if a["first"]["doc"]:
            ws2.cell(row=r, column=5,
                     value=f'=HYPERLINK($J$1&H{r}&"/detail/"&I{r},"p."&I{r})')
            ws2.cell(row=r, column=8, value=a["first"]["doc"])
            ws2.cell(row=r, column=9, value=a["first"]["page"])
        rows_by_play[play].append(r)
        r += 1
    add_dv(ws2, "F", rows_by_play)

    # ---- README -----------------------------------------------------------
    ws0 = wb.create_sheet("README", 0)
    ws0.column_dimensions["A"].width = 110
    for i, line in enumerate(
        [
            "Speaker-who review — YiDraCor (2026-07-21)",
            "",
            "Tab 'Missing who': speaker spans with NO who at all. Please pick the right role from the dropdown in column F.",
            "Tab 'Mappings to check': every distinct speaker label whose assigned who is NOT obviously its own name "
            "(pronouns, epithets, nicknames, odd spellings). If the current who (column C) is right, leave column F EMPTY. "
            "If wrong, pick the right role from the dropdown in column F.",
            "",
            "The dropdown lists the roles of that row's play. Special options: '?? unsure', 'XX not a speaker', "
            "'++ new role' — for these please add a word in Notes.",
            "'Count' = how many speeches in the play use this exact label→who mapping; your answer applies to all of them.",
            "'Example link' opens the first page in Transkribus where the mapping occurs.",
            "",
            "Not shown here: 385 mappings where the label is simply the role's own name (exact/near-exact match) — "
            "kept in the repo (speaker_who_mappings_trivial TSV); tell Sinai if you want them too.",
        ]
    ):
        ws0.cell(row=i + 1, column=1, value=line)
    ws0["A1"].font = hdr_font

    wb.save(out_xlsx)
    return len(check), len(trivial), trivial


def main():
    occurrences, roles_by_play = collect()
    today = dt.date.today().isoformat()
    out_dir = DATA / "review"
    out_dir.mkdir(exist_ok=True)
    out_xlsx = out_dir / f"speaker_who_review_{today}.xlsx"
    n_missing = sum(1 for o in occurrences if not o["xmlid"])
    n_check, n_trivial, trivial = build_workbook(occurrences, roles_by_play, out_xlsx)

    out_tsv = out_dir / f"speaker_who_occurrences_{today}.tsv"
    with open(out_tsv, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["play", "doc", "page", "file", "line_id", "label", "xmlid", "line_text"])
        for o in occurrences:
            w.writerow([o["play"], o["doc"], o["page"], o["file"], o["line_id"],
                        o["label"], o["xmlid"], o["line_text"]])

    triv_tsv = out_dir / f"speaker_who_mappings_trivial_{today}.tsv"
    with open(triv_tsv, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["play", "label", "xmlid", "count", "pages"])
        for play, label, xmlid, a in sorted(trivial):
            w.writerow([play, label, xmlid, a["n"], ",".join(map(str, sorted(a["pages"])))])

    print(f"{len(occurrences)} speaker spans; {n_missing} missing xmlid; "
          f"{n_check} mappings to check; {n_trivial} trivial mappings")
    print(f"wrote {out_xlsx}")
    print(f"wrote {out_tsv}")
    print(f"wrote {triv_tsv}")


if __name__ == "__main__":
    main()
