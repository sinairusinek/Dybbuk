"""Convert ZylbercweigPeople xlsx review artifacts into TSVs.

Outputs (next to this script):
  - people_db.tsv               — extracted "DiJeSt DB" / Zylbercweig-people target rows
                                  (DB-ID, hebrew_name, english_name, source_url_fragment)
  - people_alignment_review.tsv — RA alignment + duplication decisions per subject-entry
                                  (xml_id, heading, volume, DB-ID, same_person, action)
  - people_index_unmatched.tsv  — index names not matched to any entry heading
                                  ("לא מזוהים מהאינדקס" columns, per volume)
  - mention_validations_full.tsv     — full mention -> heading mapping w/ occurrence counts
  - mention_validations_initials.tsv — initial-form -> heading (e.g. "ב. גאָרין")
  - mention_validations_surnames.tsv — bare surname -> heading w/ disambiguation
  - credited_persons.tsv             — byline credits ("מ. ע. פֿון X") + where they appear

These are read-only conversions. They form the ground-truth pool for the
people-matcher drafter + holdout tests (analogous to Tests 1/2 for orgs).
"""
from __future__ import annotations
import csv, re
from pathlib import Path
import openpyxl

HERE = Path(__file__).parent
PEOPLE_XLSX = HERE.parent / "ZylbercweigPeople"
JSON_REVIEW = PEOPLE_XLSX / "Json Review (8).xlsx"
NEW_PEOPLE_XLSX = PEOPLE_XLSX / "ZylberberzweigPeople (8).xlsx"
REPORT_TSV = PEOPLE_XLSX / "ZylbereportPeople.tsv"
CREDITED = PEOPLE_XLSX / "credited persons.xlsx"
NAME_VALID = PEOPLE_XLSX / "extracted names validations.xlsx"
ALIGNMENT3_XLSX = PEOPLE_XLSX / "4ZylbercweigAlignment3 (2).xlsx"

DB_RE = re.compile(r"^\s*(\d+)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*$")


def _norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() == "none":
        return ""
    return s.replace("\t", " ").replace("\n", " ")


DB_FIELDS = [
    "db_id", "hebname", "english", "alternative_name", "gender",
    "date_born", "date_died", "born_in", "died_in",
    "external_source_id", "external_source",
    "author_of", "editor_of", "mentioned_in",
    "professional_role_org", "professional_role_perf", "character_perf",
    "created_expressions", "name_variants",
    "probably_not_zylbercweig", "source", "raw",
]


def _hebname_truncated(s: str) -> bool:
    """DBKey rows are upstream-truncated at first comma inside parens, e.g.
    'סאמי אוריך (שמאי' (open paren, no close). Detect that pattern so we can
    prefer Report's fuller hebrew when available."""
    if not s:
        return False
    return s.count("(") > s.count(")")


def parse_db_keys(new_xlsx_path: Path, report_tsv_path: Path):
    """Build the people DB from two sources:
      * DBKey sheet in ZylberberzweigPeople (8).xlsx — authoritative id universe
        (~3,476 well-formed ids), format 'English|id|Hebrew'.
      * ZylbereportPeople.tsv — DiJeSt biographical export (~2,996 ids, all a
        subset of DBKey).
    Returns dict id -> row, contradictions list.
    """
    import csv as _csv
    _csv.field_size_limit(10 ** 8)

    db = {}
    contradictions = []

    # 1) DBKey sheet → spine
    wb = openpyxl.load_workbook(new_xlsx_path, data_only=True)
    if "DBKey" not in wb.sheetnames:
        raise RuntimeError(f"DBKey sheet not found in {new_xlsx_path}")
    ws = wb["DBKey"]
    malformed_dbkey = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = _norm(row[0] if row else "")
        if not v:
            continue
        parts = [p.strip() for p in v.split("|")]
        if len(parts) == 3 and parts[1].isdigit():
            eng, dbid, heb = parts[0], parts[1], parts[2]
            db[dbid] = {f: "" for f in DB_FIELDS}
            db[dbid].update({
                "db_id": dbid, "english": eng, "hebname": heb,
                "source": "DBKey", "raw": v,
            })
        else:
            malformed_dbkey += 1
    wb.close()

    # 2) Report TSV → enrichment
    enriched = 0
    only_report = 0
    with open(report_tsv_path, encoding="utf-8") as fp:
        reader = _csv.DictReader(fp, delimiter="\t")
        for r in reader:
            dbid = _norm(r.get("Id"))
            if not dbid:
                continue
            r_eng = _norm(r.get("English"))
            r_heb = _norm(r.get("Hebrew"))
            issue = _norm(r.get("issue"))

            existing = db.get(dbid)
            if not existing:
                # report rows not in DBKey — shouldn't happen given current
                # data (overlap = 100% of report) but handle defensively.
                only_report += 1
                db[dbid] = {f: "" for f in DB_FIELDS}
                db[dbid].update({"db_id": dbid, "english": r_eng, "hebname": r_heb,
                                 "source": "Report-only", "raw": _norm(r.get("key"))})
                existing = db[dbid]
            else:
                # contradictions
                if r_eng and existing["english"] and r_eng != existing["english"]:
                    contradictions.append((dbid, "english", existing["english"], r_eng))
                # hebrew: prefer Report when DBKey is truncated
                if r_heb and existing["hebname"] and r_heb != existing["hebname"]:
                    if _hebname_truncated(existing["hebname"]):
                        # silent fix — known upstream truncation
                        existing["hebname"] = r_heb
                    else:
                        contradictions.append((dbid, "hebname", existing["hebname"], r_heb))
                elif r_heb and not existing["hebname"]:
                    existing["hebname"] = r_heb
                if r_eng and not existing["english"]:
                    existing["english"] = r_eng
                existing["source"] = "DBKey+Report"
                enriched += 1

            existing.update({
                "alternative_name": _norm(r.get("Alternative name")),
                "gender": _norm(r.get("Gender")),
                "date_born": _norm(r.get("Date Born")),
                "date_died": _norm(r.get("Date Died")),
                "born_in": _norm(r.get("Born in")),
                "died_in": _norm(r.get("Died in")),
                "external_source_id": _norm(r.get("External Source Id")),
                "external_source": _norm(r.get("External Source")),
                "author_of": _norm(r.get("Author of Work")),
                "editor_of": _norm(r.get("Editor of Edition")),
                "mentioned_in": _norm(r.get("Mentioned in Edition")),
                "professional_role_org": _norm(r.get("Professional Role (Organization)")),
                "professional_role_perf": _norm(r.get("Professional Role (Performance Event)")),
                "character_perf": _norm(r.get("Character (Performance Event)")),
                "created_expressions": _norm(r.get("Created expression(s)")),
                "name_variants": _norm(r.get("Name(s)")),
                "probably_not_zylbercweig": "1" if "probably not zylbercweig" in issue.lower() else "",
            })

    print(f"  DBKey: {len(db) - only_report - 0} ids ({malformed_dbkey} malformed skipped)")
    print(f"  Report enriched: {enriched}, Report-only: {only_report}")
    print(f"  Contradictions logged: {len(contradictions)}")
    return db, contradictions


def _strip_vol_prefix(xml_id: str) -> str:
    """Alignment3 uses 'N-facs_...' while everything else uses 'facs_...'."""
    if xml_id and len(xml_id) > 2 and xml_id[1] == "-" and xml_id[0].isdigit():
        return xml_id[2:]
    return xml_id


def parse_alignment3(alignment3_path: Path) -> tuple[list[dict], list[dict]]:
    """Returns (alignment_rows, db_correction_rows).

    'extracted from editions' sheet → subject-entry xml_id → db_id mapping
    (later round than the alignment+Alignment 2 sheets in Json Review).

    'Sheet1' → manual corrections to existing DB rows (kept aside, NOT applied
    to people_db.tsv automatically).
    """
    if not alignment3_path.exists():
        return [], []
    wb = openpyxl.load_workbook(alignment3_path, data_only=True)
    align_rows = []
    if "extracted from editions" in wb.sheetnames:
        ws = wb["extracted from editions"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            xid_raw = _norm(r[0] if r and len(r) > 0 else "")
            db_raw  = _norm(r[1] if r and len(r) > 1 else "")
            heading = _norm(r[2] if r and len(r) > 2 else "")
            if not xid_raw:
                continue
            xid = _strip_vol_prefix(xid_raw)
            db_id = ""
            if db_raw:
                # 'fromalignment2' is a bare int already
                db_id = db_raw.split(".")[0] if db_raw.replace(".", "").isdigit() else db_raw
            align_rows.append({
                "source_sheet": "alignment3",
                "volume": "",
                "xml_id": xid,
                "chunk_number": "",
                "heading": heading,
                "task": "",
                "action": "",
                "same_person": "",
                "decision": "",
                "db_id": db_id,
                "db_hebname": "",
                "db_english": "",
                "db_id_raw": db_raw,
                "duplication_study": "",
                "notes": "",
                "additional_pages": "",
                "additional_volume": "",
            })

    db_correction_rows = []
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        header = [_norm(c.value) for c in ws[1]]
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = {h: _norm(v) for h, v in zip(header, r)}
            did = d.get("Id", "")
            if not did:
                continue
            # Id arrives as float-ish; normalize to int string
            did = did.split(".")[0] if did.replace(".", "").isdigit() else did
            db_correction_rows.append({
                "db_id": did,
                "db_hebrew": d.get("Hebrew in DB", ""),
                "db_english": d.get("English", ""),
                "name_in_volume": d.get("שם כפי שמופיע בכרך", ""),
                "additional_appearance": d.get("מופע נוסף אם יש", ""),
                "notes": d.get("הערות", ""),
                "where_to_fix": d.get("איפה לתקן", ""),
                "correct_form": d.get("הערך הנכון (איך צריך להיות)", ""),
                "gender": d.get("Gender", ""),
                "external_source_id": d.get("External Source Id", ""),
                "external_source": d.get("External Source", ""),
                "name_variants": d.get("Name(s)", ""),
            })

    wb.close()
    return align_rows, db_correction_rows


def parse_alignment_review(json_review_path: Path, new_xlsx_path: Path = None):
    """Walk per-volume sheets + Duplication Check + alignment sheet, emit per-row
    RA decisions tied to xml_id. We don't try to interpret every column — we keep
    the fields most likely to be ground-truth signal.
    """
    wb = openpyxl.load_workbook(json_review_path, read_only=True, data_only=True)
    rows = []
    # Duplication Check
    if "Duplication Check" in wb.sheetnames:
        ws = wb["Duplication Check"]
        header = None
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [_norm(c) for c in r]
                continue
            d = {h: _norm(v) for h, v in zip(header, r)}
            if not d.get("_ - xml:id"):
                continue
            rows.append({
                "source_sheet": "Duplication Check",
                "volume": d.get("vol", ""),
                "xml_id": d.get("_ - xml:id", ""),
                "chunk_number": d.get("_ - chunk_number", ""),
                "heading": d.get("_ - heading", ""),
                "task": d.get("task", ""),
                "action": d.get("action", ""),
                "same_person": d.get("Same person?", ""),
                "db_id_raw": d.get("DB-ID", ""),
                "duplication_study": d.get("duplication study", ""),
                "notes": d.get("הערות", ""),
                "additional_pages": d.get("additional pages", ""),
                "additional_volume": d.get("additional volume", ""),
                "decision": "",  # not present in this sheet
            })
    # main alignment sheet
    if "alignment" in wb.sheetnames:
        ws = wb["alignment"]
        header = None
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [_norm(c) for c in r]
                continue
            d = {h: _norm(v) for h, v in zip(header, r)}
            if not d.get("_ - xml:id"):
                continue
            rows.append({
                "source_sheet": "alignment",
                "volume": d.get("vol", ""),
                "xml_id": d.get("_ - xml:id", ""),
                "chunk_number": d.get("_ - chunk_number", ""),
                "heading": d.get("_ - heading", ""),
                "task": "",
                "action": d.get("action", ""),
                "same_person": "",
                "db_id_raw": d.get("DB-ID", ""),
                "duplication_study": "",
                "notes": d.get("הערות", ""),
                "additional_pages": d.get("additional pages", ""),
                "additional_volume": d.get("additional volume", ""),
                "decision": d.get("Column", ""),  # used in this sheet for role/cat
            })
    wb.close()
    # merge Alignment 2 from the newer xlsx
    if new_xlsx_path and new_xlsx_path.exists():
        wb2 = openpyxl.load_workbook(new_xlsx_path, data_only=True)
        if "Alignment 2" in wb2.sheetnames:
            ws = wb2["Alignment 2"]
            header = None
            for i, r in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    header = [_norm(c) for c in r]
                    continue
                d = {h: _norm(v) for h, v in zip(header, r)}
                if not d.get("_ - xml:id"):
                    continue
                # the DB id col is named verbosely; pick whatever has 'DBKey' in it
                db_raw = ""
                for k, v in d.items():
                    if k and "DBKey" in k:
                        db_raw = v
                        break
                rows.append({
                    "source_sheet": "Alignment 2",
                    "volume": d.get("vol", ""),
                    "xml_id": d.get("_ - xml:id", ""),
                    "chunk_number": d.get("_ - chunk_number", ""),
                    "heading": d.get("_ - heading", ""),
                    "task": "",
                    "action": d.get("action", ""),
                    "same_person": "",
                    "db_id_raw": db_raw,
                    "duplication_study": "",
                    "notes": d.get("הערות", ""),
                    "additional_pages": "",
                    "additional_volume": "",
                    "decision": d.get("Resp.", ""),
                })
        wb2.close()
    # parse db_id_raw into a numeric id if possible
    for row in rows:
        raw = row["db_id_raw"]
        m = DB_RE.match(raw) if raw else None
        if m:
            row["db_id"] = m.group(1)
            row["db_hebname"] = m.group(2)
            row["db_english"] = m.group(3)
        else:
            # sometimes raw is just a float-looking id
            row["db_id"] = raw.split(".")[0] if raw and raw.replace(".", "").isdigit() else ""
            row["db_hebname"] = ""
            row["db_english"] = ""
    return rows


def parse_index_unmatched(json_review_path: Path):
    """Per-volume 'align' sheets list index names that didn't match an entry heading."""
    wb = openpyxl.load_workbook(json_review_path, read_only=True, data_only=True)
    rows = []
    align_sheets = [s for s in wb.sheetnames if s.lower().endswith("align") or "align" in s.lower() and s != "alignment"]
    for sname in align_sheets:
        ws = wb[sname]
        header = None
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [_norm(c) for c in r]
                continue
            d = {h: _norm(v) for h, v in zip(header, r)}
            # the "unmatched" payload sits in an explicit column or trailing columns
            unmatched = d.get("לא מזוהים מהאינדקס", "")
            heading = d.get("_ - heading", "")
            vol = d.get("vol", "")
            xml_id = d.get("_ - xml:id", "")
            # also pick up any non-empty 'Column...' values which are extra index names
            extras = []
            for k, v in d.items():
                if v and (k.startswith("Column") or k == "Column 1 2 in Zylbercweig people"):
                    extras.append(v)
            if not (unmatched or heading or extras):
                continue
            rows.append({
                "source_sheet": sname,
                "volume": vol,
                "xml_id": xml_id,
                "heading": heading,
                "index_unmatched": unmatched,
                "extras": " || ".join(extras),
            })
    wb.close()
    return rows


def parse_name_validations(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = {"full": [], "initials": [], "surnames": []}

    def harvest(sheet_name, key_col, freq_col, heading_col, alt_col=None, notes_col=None):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        header = None
        rows = []
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [_norm(c) for c in r]
                continue
            d = {h: _norm(v) for h, v in zip(header, r)}
            mn = d.get(key_col, "")
            if not mn:
                continue
            rows.append({
                "mention": mn,
                "occurrences": d.get(freq_col, ""),
                "as_heading": d.get(heading_col, ""),
                "alternative_heading": d.get(alt_col, "") if alt_col else "",
                "notes": d.get(notes_col, "") if notes_col else "",
            })
        return rows

    out["full"] = harvest("validate extracted names", "mentioned name", "occurances as mentioned", "as entry heading", notes_col="הערות")
    out["initials"] = harvest("initials", "initials", "occurences", "as heading", alt_col="alternative heading", notes_col="הערות")
    out["surnames"] = harvest("surnames only", "Single-word", "occurences", "as heading", alt_col="alternative heading for disambiguation", notes_col="הערות ושאלות")
    wb.close()
    return out


def parse_credited(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    header = None
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [_norm(c) for c in r]
            continue
        d = {h: _norm(v) for h, v in zip(header, r)}
        if not d.get("Credit"):
            continue
        rows.append({
            "credit": d.get("Credit", ""),
            "occurrences": d.get("where to find it", ""),
            "clustered": d.get("clustered", ""),
            "full_name": d.get("add full name here", ""),
            "comments": d.get("comments if there are any", ""),
            "qid": d.get("Qid", ""),
        })
    wb.close()
    return rows


def write_tsv(path: Path, rows, fieldnames):
    with open(path, "w", encoding="utf-8", newline="") as fp:
        w = csv.DictWriter(fp, fieldnames=fieldnames, delimiter="\t", extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"wrote {path} ({len(rows)} rows)")


def main():
    db, contradictions = parse_db_keys(NEW_PEOPLE_XLSX, REPORT_TSV)
    # sort by numeric id for stable output
    rows_sorted = sorted(db.values(), key=lambda r: int(r["db_id"]) if r["db_id"].isdigit() else 10**9)
    write_tsv(HERE / "people_db.tsv", rows_sorted, DB_FIELDS)

    if contradictions:
        with open(HERE / "people_db_contradictions.tsv", "w", encoding="utf-8") as fp:
            fp.write("db_id\tfield\tdbkey_value\treport_value\n")
            for did, field, a, b in contradictions:
                fp.write(f"{did}\t{field}\t{a}\t{b}\n")
        print(f"wrote {HERE / 'people_db_contradictions.tsv'} ({len(contradictions)} rows)")

    align = parse_alignment_review(JSON_REVIEW, NEW_PEOPLE_XLSX)

    # Alignment3 (later round) — merge in. Disagreements between old alignment
    # rounds and alignment3 are dumped to a separate TSV for review; we do NOT
    # silently pick a winner.
    align3, db_corrections = parse_alignment3(ALIGNMENT3_XLSX)
    if align3:
        # Build the first non-empty db_id per xml_id from the older rounds
        old_by_xid: dict[str, str] = {}
        for row in align:
            xid = row.get("xml_id", "")
            did = row.get("db_id", "")
            if xid and did and xid not in old_by_xid:
                old_by_xid[xid] = did
        disagreements = []
        n_agree = n_only_new = n_disagree = 0
        for row in align3:
            xid = row.get("xml_id", "")
            new_did = row.get("db_id", "")
            if not (xid and new_did):
                continue
            prev = old_by_xid.get(xid, "")
            if not prev:
                n_only_new += 1
            elif prev == new_did:
                n_agree += 1
            else:
                n_disagree += 1
                disagreements.append({
                    "xml_id": xid,
                    "heading": row.get("heading", ""),
                    "older_db_id": prev,
                    "alignment3_db_id": new_did,
                })
        align.extend(align3)
        print(f"\nalignment3 merge: agree={n_agree}, only_new={n_only_new}, disagree={n_disagree}")
        if disagreements:
            with open(HERE / "alignment_disagreements.tsv", "w", encoding="utf-8", newline="") as fp:
                w = csv.DictWriter(
                    fp,
                    fieldnames=["xml_id", "heading", "older_db_id", "alignment3_db_id"],
                    delimiter="\t",
                )
                w.writeheader()
                w.writerows(disagreements)
            print(f"wrote {HERE / 'alignment_disagreements.tsv'} ({len(disagreements)} rows)")

    if db_corrections:
        with open(HERE / "db_correction_annotations.tsv", "w", encoding="utf-8", newline="") as fp:
            w = csv.DictWriter(
                fp,
                fieldnames=["db_id", "db_hebrew", "db_english", "name_in_volume",
                            "additional_appearance", "notes", "where_to_fix",
                            "correct_form", "gender", "external_source_id",
                            "external_source", "name_variants"],
                delimiter="\t",
            )
            w.writeheader()
            w.writerows(db_corrections)
        print(f"wrote {HERE / 'db_correction_annotations.tsv'} ({len(db_corrections)} rows)")
    write_tsv(HERE / "people_alignment_review.tsv", align,
              ["source_sheet", "volume", "xml_id", "chunk_number", "heading",
               "task", "action", "same_person", "decision",
               "db_id", "db_hebname", "db_english", "db_id_raw",
               "duplication_study", "notes", "additional_pages", "additional_volume"])

    unmatched = parse_index_unmatched(JSON_REVIEW)
    write_tsv(HERE / "people_index_unmatched.tsv", unmatched,
              ["source_sheet", "volume", "xml_id", "heading", "index_unmatched", "extras"])

    nv = parse_name_validations(NAME_VALID)
    write_tsv(HERE / "mention_validations_full.tsv", nv["full"],
              ["mention", "occurrences", "as_heading", "alternative_heading", "notes"])
    write_tsv(HERE / "mention_validations_initials.tsv", nv["initials"],
              ["mention", "occurrences", "as_heading", "alternative_heading", "notes"])
    write_tsv(HERE / "mention_validations_surnames.tsv", nv["surnames"],
              ["mention", "occurrences", "as_heading", "alternative_heading", "notes"])

    credited = parse_credited(CREDITED)
    write_tsv(HERE / "credited_persons.tsv", credited,
              ["credit", "occurrences", "clustered", "full_name", "comments", "qid"])


if __name__ == "__main__":
    main()
