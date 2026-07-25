"""C1 — Normalize the held-out curated catalogue into evaluation tables.

Reads the DybbukCatalogue xlsx (works, hafakot/production, role sheets for
Lateiner + Hurwitz), the latest PerformanceEvents_Report_*.xlsx, and
editions.json `productions`. These were built largely from EXTERNAL sources
(newspapers, Sieger, posters, archives) — they are an evaluation reference
ONLY and are never fed into extraction.

Outputs -> eval/:
  eval_reference_works.tsv        one row per catalogued expression
  eval_reference_productions.tsv  one row per premiere/production/show event
  eval_reference_roles.tsv        one row per (play, person, role)

Usage:
    python3.11 build_eval_reference.py            # dry-run stats
    python3.11 build_eval_reference.py --execute
"""
from __future__ import annotations

import argparse
import json

import openpyxl

import plays_common as pc

CATALOGUE_XLSX = pc.EDITION_METADATA_DIR / "DybbukCatalogue May2024.xlsx"

WORKS_FIELDS = ["ref_id", "source_sheet", "expression_id", "title_en", "title_yi",
                "author", "genre", "certainty", "title_segments_norm"]
PROD_FIELDS = ["ref_id", "source_sheet", "play_key", "title_yi", "event_type",
               "year", "date", "place", "theatre", "source_citation", "notes",
               "title_segments_norm"]
ROLE_FIELDS = ["ref_id", "source_sheet", "play", "person", "person_key", "role",
               "character", "context", "source_citation", "title_segments_norm"]


def sheet_rows(wb, name: str) -> list[dict]:
    if name not in wb.sheetnames:
        print(f"  ! sheet missing: {name}")
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else f"col{i}"
              for i, h in enumerate(rows[0])]
    out = []
    for r in rows[1:]:
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        out.append({header[i]: ("" if v is None else str(v).strip())
                    for i, v in enumerate(r) if i < len(header)})
    return out


def col(row: dict, *names: str) -> str:
    for n in names:
        for k, v in row.items():
            if k.lower().replace(" ", "") == n.lower().replace(" ", ""):
                return v
    return ""


def clean_year(v: str) -> str:
    v = v.split(".")[0]
    return v if v.isdigit() and len(v) == 4 else ""


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--execute", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(CATALOGUE_XLSX, read_only=True, data_only=True)
    print(f"catalogue sheets: {len(wb.sheetnames)}")

    works, prods, roles = [], [], []

    for sheet, author in (("Lateiner Plays", "Lateiner"), ("Hurwitz Plays", "Hurwitz")):
        for r in sheet_rows(wb, sheet):
            title_en = col(r, "English Name")
            title_yi = col(r, "Yiddish Name")
            if not (title_en or title_yi):
                continue
            works.append({
                "source_sheet": sheet,
                "expression_id": col(r, "Expression ID").split(".")[0],
                "title_en": title_en, "title_yi": title_yi,
                "author": col(r, "author") or author,
                "genre": col(r, "Genre"), "certainty": col(r, "certainty"),
                "title_segments_norm": "|".join(pc.title_segments(title_yi)),
            })

    for sheet in ("Lateiner hafakot", "Hurwitz hafakot"):
        for r in sheet_rows(wb, sheet):
            key = col(r, "Play KEY")
            if not key:
                continue
            other = col(r, "playNamefromOtherSources")
            prods.append({
                "source_sheet": sheet, "play_key": key,
                "title_yi": other if any("א" <= c <= "ת" for c in other) else "",
                "event_type": col(r, "Type"),
                "year": clean_year(col(r, "Year")),
                "date": col(r, "innacurate or exact date", "date"),
                "place": col(r, "PremierePlace"),
                "theatre": col(r, "Theatre"),
                "source_citation": col(r, "source", "dating Source"),
                "notes": col(r, "Notes"),
                "title_segments_norm": "|".join(pc.title_segments(other)),
            })

    # DB export of performance events (picked by latest mtime, like the
    # YiDraCor build script does)
    pe_files = sorted(pc.EDITION_METADATA_DIR.glob("PerformanceEvents_Report_*.xlsx"))
    if pe_files:
        wb2 = openpyxl.load_workbook(pe_files[-1], read_only=True, data_only=True)
        name = next((s for s in wb2.sheetnames if "performance" in s.lower()),
                    wb2.sheetnames[0])
        for r in sheet_rows(wb2, name):
            title_yi = col(r, "Yiddish Title")
            prods.append({
                "source_sheet": f"PerformanceEvents:{pe_files[-1].name}",
                "play_key": col(r, "Roman Title"),
                "title_yi": title_yi,
                "event_type": col(r, "Event Type"),
                "year": clean_year(col(r, "Date")[-4:] if col(r, "Date") else ""),
                "date": col(r, "Date"),
                "place": "", "theatre": col(r, "Held at"),
                "source_citation": col(r, "source_catalogue", "source"),
                "notes": " | ".join(x for x in (col(r, "Actor (Character)"),
                                                col(r, "Person (Professional Role)")) if x),
                "title_segments_norm": "|".join(pc.title_segments(title_yi)),
            })
            actor = col(r, "Actor (Character)")
            person_role = col(r, "Person (Professional Role)")
            for blob, kind in ((actor, "actor"), (person_role, "role")):
                if blob:
                    roles.append({
                        "source_sheet": f"PerformanceEvents:{pe_files[-1].name}",
                        "play": col(r, "Roman Title"), "person": blob,
                        "person_key": "", "role": kind, "character": "",
                        "context": col(r, "Event Type") + " " + col(r, "Date"),
                        "source_citation": col(r, "source_catalogue", "source"),
                        "title_segments_norm": "|".join(pc.title_segments(title_yi)),
                    })

    for sheet in ("ProfessionalRoles- Lateiner", "playRolesLateiner", "PlayRolesHurwitz"):
        for r in sheet_rows(wb, sheet):
            play = col(r, "Play", "Play Key", "play")
            person = col(r, "PersonName", "Person", "name", "PersonNameEnglishspelling")
            if not play and not person:
                continue
            roles.append({
                "source_sheet": sheet, "play": play, "person": person,
                "person_key": col(r, "PersonKey"),
                "role": col(r, "Role"), "character": col(r, "character", "Character"),
                "context": col(r, "context"), "source_citation": col(r, "source"),
                "title_segments_norm": "|".join(pc.title_segments(play)),
            })

    # editions.json productions (catalogue layer embedded in the build output)
    ed_json = json.loads((pc.YIDRACOR_DATA / "editions.json").read_text(encoding="utf-8"))
    for e in ed_json.get("editions", []):
        for ev in e.get("performance_events") or []:
            title_yi = str(ev.get("yiddish_title") or "")
            prods.append({
                "source_sheet": "editions.json:performance_events",
                "play_key": str(ev.get("roman_title") or ""),
                "title_yi": title_yi,
                "event_type": str(ev.get("event_type") or ""),
                "year": str(ev.get("year") or "").split(".")[0],
                "date": str(ev.get("date") or ""),
                "place": "", "theatre": str(ev.get("venue") or ""),
                "source_citation": str(ev.get("source_catalogue") or ""),
                "notes": str(ev.get("notes") or ""),
                "title_segments_norm": "|".join(pc.title_segments(title_yi)),
            })

    for i, r in enumerate(works, 1):
        r["ref_id"] = f"W-{i:04d}"
    for i, r in enumerate(prods, 1):
        r["ref_id"] = f"PR-{i:04d}"
    for i, r in enumerate(roles, 1):
        r["ref_id"] = f"R-{i:04d}"

    print(f"works: {len(works)}  productions/events: {len(prods)}  roles: {len(roles)}")
    if not args.execute:
        print("dry-run — pass --execute to write eval/ tables")
        return
    pc.write_tsv(pc.EVAL_DIR / "eval_reference_works.tsv", works, WORKS_FIELDS)
    pc.write_tsv(pc.EVAL_DIR / "eval_reference_productions.tsv", prods, PROD_FIELDS)
    pc.write_tsv(pc.EVAL_DIR / "eval_reference_roles.tsv", roles, ROLE_FIELDS)
    print(f"wrote eval/ reference tables")


if __name__ == "__main__":
    main()
