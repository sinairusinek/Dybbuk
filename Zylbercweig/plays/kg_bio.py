"""KG bio layer — lexicon subjects + birth/death/burial places.

Step 0 (entry index) + Step 1 (biography layer) of the Colab-extraction ->
KG integration.  Called from build_kg.py after the plays layer so that every
layer shares one Graph (node ids, mint sequences, adjudication).

Sources
  people/people_extracted.tsv        one row per lexicon subject-entry
                                     (flatten_volumes.py over the Colab "III"
                                     JSONs): heading, entry_type, dates, places
  people/person_hub.tsv              entry_person_id -> people_db db_id
                                     (RA/human alignment, consolidated)
  zibn-shtern/.../toponyms_attestations.csv
                                     per-entry birth/death/burial place
                                     attestations with link_status + QID
                                     (build_unified_toponyms.py)

Outputs (via build_kg.py --execute)
  kg/entry_index.tsv                 the join table every later layer uses:
                                     entry_key, person_id, db_id, node_id ...
  person / place nodes + born_in / died_in / buried_in edges in kg/nodes.tsv
  and kg/edges.tsv, stamped source_layer=bio.

Node-id policy (same as the plays layer): a subject aligned to people_db is
person:<db_id>; an unaligned subject is person_entry:<person_id>.  Places are
place:<QID> when the attestation is linked/needs_review, else minted
place:UPL-nnnn keyed on the normalized surface (shared with the plays layer so
the same unresolved toponym collapses to one node).
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict

import plays_common as pc

PERSON_HUB_TSV = pc.PEOPLE_DIR / "person_hub.tsv"
PEOPLE_EXTRACTED_TSV = pc.PEOPLE_DIR / "people_extracted.tsv"
ATTESTATIONS_CSV = pc.ZIBN_WORKING / "toponyms_attestations.csv"
ENTRY_INDEX_TSV = pc.KG_DIR / "entry_index.tsv"

ENTRY_INDEX_FIELDS = ["entry_key", "person_id", "volume", "xml_id", "heading",
                      "entry_type", "hub_id", "db_id", "node_id", "align_evidence"]

CONTEXT_EDGE = {"birth": "born_in", "death": "died_in", "burial": "buried_in"}
LINK_CONF = {"linked": "high", "needs_review": "medium"}
EXTRACTION_MODEL = "colab_extraction_III"


# ---------------------------------------------------------------- step 0
def build_entry_index() -> dict[str, dict]:
    """person_id -> entry-index row. Alignment authority = person_hub.tsv."""
    hub_of: dict[str, tuple[str, str, str]] = {}
    for h in pc.read_tsv(PERSON_HUB_TSV):
        db_ids = [d for d in (h.get("db_ids") or "").split("|") if d]
        if len(db_ids) > 1:  # multi_db hubs are unresolved — don't pick one
            db_ids = []
        for pid in (h.get("entry_person_ids") or "").split("|"):
            if pid:
                hub_of[pid] = (h["hub_id"], db_ids[0] if db_ids else "",
                               h.get("evidence") or "")
    index: dict[str, dict] = {}
    for r in pc.read_tsv(PEOPLE_EXTRACTED_TSV):
        pid = r["person_id"]
        hub_id, db_id, evidence = hub_of.get(pid, ("", "", ""))
        index[pid] = {
            "entry_key": f"{r['volume']}-{r['xml_id']}",
            "person_id": pid, "volume": r["volume"], "xml_id": r["xml_id"],
            "heading": r["heading"], "entry_type": r.get("entry_type", ""),
            "hub_id": hub_id, "db_id": db_id,
            "node_id": f"person:{db_id}" if db_id else f"person_entry:{pid}",
            "align_evidence": evidence,
            # carried for the bio layer, not written to entry_index.tsv
            "_row": r,
        }
    return index


def entry_to_dbid_map(index: dict[str, dict]) -> dict[str, str]:
    return {pid: e["db_id"] for pid, e in index.items() if e["db_id"]}


# ---------------------------------------------------------------- helpers
_DATE_RE = re.compile(r"^(\d{4})(?:-(\d{2}))?(?:-(\d{2}))?$")


def norm_date(raw: str) -> tuple[str, str, str]:
    """(iso_start, iso_end, precision) for the extraction's date strings.
    Handles YYYY, YYYY-MM, YYYY-MM-DD, 'YYYY~', 'YYYY/YYYY', '189X'."""
    s = (raw or "").strip()
    if not s:
        return "", "", ""
    m = _DATE_RE.match(s)
    if m:
        y, mo, d = m.groups()
        if d:
            return s, "", "day"
        if mo:
            return s, "", "month"
        return s, "", "year"
    if re.fullmatch(r"\d{4}~", s):
        return s[:4], "", "circa"
    m = re.fullmatch(r"(\d{4})/(\d{4})", s)
    if m:
        return m.group(1), m.group(2), "range"
    m = re.fullmatch(r"(\d{3})[xX.]", s) or re.fullmatch(r"(\d{3})", s)
    if m:
        return m.group(1) + "0", m.group(1) + "9", "decade"
    m = re.fullmatch(r"(\d{2})[xX.]{2}", s)
    if m:
        return m.group(1) + "00", m.group(1) + "99", "century"
    return "", "", "unparsed"


def _load_attestations() -> dict[tuple[str, str], dict]:
    """(entry_key, context) -> best place attestation (field=place, then
    settlement).  Only person-corpus rows."""
    best: dict[tuple[str, str], dict] = {}
    rank = {"place": 0, "settlement": 1}
    with open(ATTESTATIONS_CSV, encoding="utf-8-sig") as f:
        for a in csv.DictReader(f):
            if a["source_corpus"] != "person" or a["source_field"] not in rank:
                continue
            key = (a["source_record_id"], a["context"])
            cur = best.get(key)
            if cur is None or rank[a["source_field"]] < rank[cur["source_field"]]:
                best[key] = a
    return best


# ---------------------------------------------------------------- step 4
# family_background / education as node attributes (not nodes).  The family
# fields carry the extraction's controlled values (financial: modest|rich|poor,
# religion: religious|maskilic_secular|yichus|christian, theater_connection:
# connected|amateur_lover|opposed, structure: orphan); education is free text
# that gets a coarse class list on top.  Only vol 7 was RA-reviewed.
EXTRACTION_DIR = pc.HERE.parent / "Zylbercweig_extraction"

EDU_CLASSES = [  # (class, regex over pc.norm_yiddish text — un-finalized letters)
    ("kheyder", r"\bחדר|\bחדרימ|\bמלמד"),
    ("yeshiva", r"ישיבה|בית המדרש|\bתלמוד תורה"),
    ("gymnasium", r"גימנאז"),
    ("folkshul", r"פאלקס ?שול|פאלקשול"),
    ("public_school", r"פאבליק ?סקול|פובליק ?סקול|\bסקול\b"),
    ("commercial_school", r"קאמערצ|האנדלס"),
    ("conservatory", r"קאנסערוואטאר"),
    ("university", r"אוניווערסיטעט|אוניווערזיטעט|\bקאלעדזש|פאקולטעט|מעדיצינ"),
    ("drama_school", r"דראמ[^ ]* ?(?:שול|סטודיא|קורס)|טעאטער ?(?:שול|סטודיא)|סטודיא"),
    ("music_school", r"מוזיק ?שול|געזאנג"),
    ("extern", r"עקסטערנ"),
    ("self_taught", r"זעלבסט ?(?:בילדונג|געלערנט)|אויטאדידאקט"),
    ("private_tutors", r"פריוואט|מיט לערער"),
]
_EDU = [(c, re.compile(rx)) for c, rx in EDU_CLASSES]


def education_classes(text: str) -> list[str]:
    s = pc.norm_yiddish(text)
    return [c for c, rx in _EDU if rx.search(s)]


def _load_family_background() -> dict[str, dict]:
    """entry_key -> family attrs from the IIIorg JSONs."""
    import glob, json as _json, os
    out: dict[str, dict] = {}
    for path in sorted(glob.glob(str(EXTRACTION_DIR / "*IIIorg.json"))):
        m = re.search(r"(\d+)", os.path.basename(path))
        vol = m.group(1) if m else ""
        with open(path, encoding="utf-8") as f:
            for e in _json.load(f):
                fb = e.get("family_background") or []
                if not fb:
                    continue
                d: dict = {}
                for b in fb:
                    for k, ak in (("financial", "family_financial"),
                                  ("religion", "family_religion"),
                                  ("theater_connection", "family_theater"),
                                  ("structure", "family_structure")):
                        if b.get(k) and ak not in d:
                            d[ak] = b[k]
                    if b.get("description"):
                        d.setdefault("family_description", []).append(b["description"])
                if "family_description" in d:
                    d["family_description"] = " | ".join(d["family_description"])
                out[f"{vol}-{(e.get('xml:id') or '').strip()}"] = d
    return out


# ---------------------------------------------------------------- step 1a
# add_node only fills EMPTY keys (see build_kg.Graph.add_node), so when 2+
# lexicon entries share a node_id (a person aligned to the same people_db id
# via 2+ volumes) only the FIRST entry's attrs would survive. Merge across
# entries here, before add_node is ever called for that node_id.
def _merge_scalar(vals: list[str]) -> tuple[str, list[str]]:
    """First non-empty value, plus other distinct non-empty values seen."""
    non_empty = [v for v in vals if v]
    if not non_empty:
        return "", []
    first = non_empty[0]
    alts = []
    for v in non_empty[1:]:
        if v != first and v not in alts:
            alts.append(v)
    return first, alts


def _merge_entry_attrs(entries: list[dict]) -> dict:
    """Merge per-entry attrs dicts (+ volume/span/credit/person_id) for all
    lexicon entries that resolve to the same node_id. `entries` items carry
    pid, attrs (dict), volume, span, credit."""
    if len(entries) == 1:
        return dict(entries[0]["attrs"])

    merged: dict = {}
    entry_list = [
        {k: v for k, v in (
            ("volume", en["volume"]), ("span", en["span"]),
            ("credit", en["credit"]), ("person_id", en["pid"]),
        ) if v}
        for en in entries
    ]
    merged["entries"] = entry_list

    excluded = {"volume", "span", "credit"}
    all_keys: list[str] = []
    for en in entries:
        for k in en["attrs"]:
            if k not in excluded and k not in all_keys:
                all_keys.append(k)

    for k in all_keys:
        vals = [en["attrs"].get(k, "") for en in entries]
        if k == "education_classes":
            union: list[str] = []
            for v in vals:
                for c in (v or []):
                    if c not in union:
                        union.append(c)
            if union:
                merged[k] = union
        elif k in ("education", "family_description"):
            parts: list[str] = []
            for v in vals:
                if v and v not in parts:
                    parts.append(v)
            if parts:
                merged[k] = " | ".join(parts)
        else:
            first, alts = _merge_scalar(vals)
            if first:
                merged[k] = first
            if alts:
                merged[f"{k}_alt"] = alts
    return merged


# ---------------------------------------------------------------- RA sheets
CEMETERIES_XLSX = EXTRACTION_DIR / "Zylbercweig-Extraction2026-02-05cemeteries-tsv.xlsx"
JSON_REVIEW_XLSX = EXTRACTION_DIR / "Json Review.xlsx"


def _load_cemeteries() -> dict[str, tuple[str, str]]:
    """entry_key -> (cemetery QID, find-a-grave url) from the RA burial sheet
    (2026-02-05; 167 QIDs). Used to fill burial places the toponym spine left
    unlinked and to confirm needs_review ones; a disagreement with a LINKED
    attestation is recorded, not applied."""
    try:
        import openpyxl
    except ImportError:
        return {}
    if not CEMETERIES_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(CEMETERIES_XLSX, read_only=True)
    rows = wb.worksheets[0].iter_rows(values_only=True)
    hdr = [str(h).strip() for h in next(rows)]
    i = {h: k for k, h in enumerate(hdr)}
    out = {}
    for r in rows:
        if not r or r[i["unique-id"]] is None:
            continue
        m = re.search(r"Q\d+", str(r[i["wikidata"]] or ""))
        fg = str(r[i["find a grave link"]] or "").strip() if "find a grave link" in i else ""
        out[str(r[i["unique-id"]]).strip()] = (m.group(0) if m else "", fg)
    return out


def _load_ra_final_types() -> dict[str, str]:
    """entry_key -> RA 'final type' from Json Review.xlsx per-volume sheets
    (374 verdicts; 'NOT AN ENTRY' marks non-subjects the LLM split off)."""
    try:
        import openpyxl
    except ImportError:
        return {}
    if not JSON_REVIEW_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(JSON_REVIEW_XLSX, read_only=True)
    out = {}
    for ws in wb.worksheets:
        if not ws.title.lower().startswith("volume"):
            continue
        rows = ws.iter_rows(values_only=True)
        hdr = [str(h) for h in next(rows)]
        if "final type" not in hdr or "_ - xml:id" not in hdr:
            continue
        i = {h: k for k, h in enumerate(hdr)}
        for r in rows:
            if not r or r[i["_ - xml:id"]] is None:
                continue
            f = str(r[i["final type"]] or "").strip()
            if f:
                vol = str(r[i["Volume"]]).split(".")[0]
                out[f"{vol}-{str(r[i['_ - xml:id']]).strip()}"] = f
    return out


# ---------------------------------------------------------------- step 1
def add_bio_layer(g, labels, index: dict[str, dict]) -> dict:
    """Add every lexicon subject as a person node and its place edges."""
    from build_kg import _place_label_yi, _place_label_en  # avoid import cycle

    people, _orgs, _clusters, places = labels
    att = _load_attestations()
    family = _load_family_background()
    cemeteries = _load_cemeteries()
    ra_types = _load_ra_final_types()
    stats: Counter = Counter()

    # pass 1: build per-entry attrs, grouped by node_id (multiple lexicon
    # entries can share a node_id when they align to the same people_db id)
    by_node: dict[str, list[dict]] = defaultdict(list)
    for pid, e in index.items():
        r = e["_row"]
        attrs = {k: v for k, v in (
            ("entry_type", r.get("entry_type", "")),
            ("volume", r["volume"]), ("span", r.get("span", "")),
            ("birth_date", r.get("birth_date", "")),
            ("death_date", r.get("death_date", "")),
            ("credit", r.get("credit", "")),
            ("education", r.get("education", "")),
        ) if v}
        if r.get("education"):
            cls = education_classes(r["education"])
            if cls:
                attrs["education_classes"] = cls
            stats["education"] += 1
        fam = family.get(e["entry_key"])
        if fam:
            attrs.update(fam)
            stats["family_background"] += 1
        ra = ra_types.get(e["entry_key"])
        if ra:
            attrs["entry_type_ra"] = ra
            stats["ra_final_type"] += 1
            if ra.upper() == "NOT AN ENTRY":
                attrs["ra_not_an_entry"] = True
                stats["ra_not_an_entry"] += 1
        cem = cemeteries.get(e["entry_key"])
        if cem and cem[1]:
            attrs["findagrave"] = cem[1]
            stats["findagrave"] += 1
        by_node[e["node_id"]].append({
            "pid": pid, "attrs": attrs,
            "volume": r["volume"], "span": r.get("span", ""),
            "credit": r.get("credit", ""),
        })
        stats["subjects"] += 1
        stats["subjects_aligned" if e["db_id"] else "subjects_unaligned"] += 1

    # pass 2: one add_node call per node_id, attrs merged across its entries
    for node_id, entries in by_node.items():
        merged_attrs = _merge_entry_attrs(entries)
        attrs_json = json.dumps(merged_attrs, ensure_ascii=False)
        pids = [en["pid"] for en in entries]
        e0 = index[pids[0]]
        r0 = e0["_row"]
        if e0["db_id"]:
            p = people.get(e0["db_id"], {})
            g.add_node(node_id, node_type="person",
                       label_yiddish=p.get("hebname") or r0["heading"],
                       label_english=p.get("english", ""),
                       ext_ref_type="people_db", ext_ref_id=e0["db_id"],
                       secondary_ids="|".join(
                           f"entry_person_id:{pid}" for pid in pids),
                       match_status="matched", source_layer="bio",
                       attrs=attrs_json)
            # add_node only fills EMPTY keys: if the plays layer already
            # upgraded this node_id earlier (person_entry: -> person:) it
            # stamped secondary_ids with a single entry_person_id, which
            # would silently block our full pid list above. Force-merge so
            # every entry's pid survives regardless of write order.
            existing = g.nodes[node_id].get("secondary_ids", "") or ""
            existing_ids = [x for x in existing.split("|") if x]
            want_ids = [f"entry_person_id:{pid}" for pid in pids]
            g.nodes[node_id]["secondary_ids"] = "|".join(
                dict.fromkeys(existing_ids + want_ids))
        else:
            g.add_node(node_id, node_type="person", label_yiddish=r0["heading"],
                       ext_ref_type="entry_person_id", ext_ref_id=pids[0],
                       match_status="unmatched", source_layer="bio",
                       attrs=attrs_json)

    # pass 3: place edges (per entry — each entry's own attestations)
    for pid, e in index.items():
        r = e["_row"]
        node_id = e["node_id"]
        for ctx, etype in CONTEXT_EDGE.items():
            surface = r.get(f"{ctx}_place_name", "")
            if not surface:
                continue
            prov = r.get(f"{ctx}_place_province", "")
            ctry = r.get(f"{ctx}_place_country", "")
            evidence = " / ".join(x for x in (surface, prov, ctry) if x)
            a = att.get((e["entry_key"], ctx))
            status = (a or {}).get("link_status", "")
            qid = (a or {}).get("qid", "")
            sheet_note = ""
            if ctx == "burial":
                cq = (cemeteries.get(e["entry_key"]) or ("", ""))[0]
                if cq:
                    if status == "linked" and qid and qid != cq:
                        sheet_note = f"cemetery_sheet_disagrees:{cq}"
                        stats["burial_sheet_conflict"] += 1
                    elif status == "needs_review" and qid == cq:
                        status = "linked"  # RA sheet confirms the spine's guess
                        stats["burial_sheet_confirms"] += 1
                    elif not qid or status not in LINK_CONF:
                        # sheet fills an unlinked / absent attestation
                        a = dict(a or {}); a.setdefault("attestation_id", "")
                        a["attestation_id"] = (a["attestation_id"] + "|" if a["attestation_id"] else "") + "cemeteries_sheet"
                        a["label_en"] = ""
                        qid, status = cq, "needs_review"
                        stats["burial_sheet_fills"] += 1
            if a and qid and status in LINK_CONF:
                pl = places.get(qid, {})
                sec = json.dumps({k: pl.get(k, "") for k in ("kima_id", "lat", "lon")},
                                 ensure_ascii=False) if pl else ""
                target = f"place:{qid}"
                g.add_node(target, node_type="place",
                           label_yiddish=_place_label_yi(pl, surface),
                           label_english=_place_label_en(pl) or a.get("label_en", ""),
                           ext_ref_type="wikidata_qid", ext_ref_id=qid,
                           secondary_ids=sec,
                           match_status="matched" if status == "linked" else "candidate",
                           source_layer="bio")
                match = "matched" if status == "linked" else "candidate"
                conf = LINK_CONF[status]
            else:
                # unlinked / misresolved / no attestation -> minted node,
                # shared mint key with the plays layer
                notes = f"attestation:{status}" if status else "no_attestation"
                if status == "misresolved" and qid:
                    notes += f" rejected:{qid}"
                target = g.resolve_surface(
                    "place", surface, "bio", (a or {}).get("attestation_id", "") or pid,
                    evidence, node_type="place", label_yiddish=surface,
                    match_status="unmatched", notes=notes, source_layer="bio")
                if not target:
                    stats[f"{etype}:not_entity"] += 1
                    continue
                if target.startswith("place:UPL"):
                    match, conf = "unmatched", "low"
                else:
                    match, conf = "matched", "high"  # adjudicated ALIGN
            date_key = {"birth": "birth_date", "death": "death_date",
                        "burial": "death_date"}[ctx]
            d0, d1, prec = norm_date(r.get(date_key, ""))
            g.add_edge(source_id=node_id, target_id=target, edge_type=etype,
                       role_detail="", character="",
                       date_start=d0, date_end=d1, date_precision=prec,
                       event_id="", production_key="",
                       provenance_person_id=pid,
                       provenance_fact_ids=(a or {}).get("attestation_id", "")
                       + (f"|{sheet_note}" if sheet_note else ""),
                       evidence_sentence=evidence,
                       extraction_model=EXTRACTION_MODEL,
                       confidence=conf, match_status=match,
                       review_status="auto", source_layer="bio")
            stats[f"{etype}:{match}"] += 1
    return dict(stats)


def write_entry_index(index: dict[str, dict]) -> None:
    pc.write_tsv(ENTRY_INDEX_TSV, list(index.values()), ENTRY_INDEX_FIELDS)
